import os
import logging
from openpyxl import Workbook, load_workbook, styles
from openpyxl.utils import get_column_letter

from src.config import Config

class Xlsx:
    def __init__(self, config: Config) -> None:
        '''表格数据操作类'''
        self.file_path = config.output_file
        if path := os.path.dirname(self.file_path):
            os.makedirs(path, exist_ok=True)

    def write(self, sheet_name: str, data: list, overwrite: bool = True) -> bool:
        '''数据写入到表格'''  
        if overwrite:
            self.wb = Workbook()
            self.wb.remove(self.wb['Sheet'])
        else:
            try:
                self.wb = load_workbook(self.file_path)
            except FileNotFoundError:
                self.wb = Workbook()
                self.wb.remove(self.wb['Sheet'])
        if len(data) <= 1:
            logging.warning('数据为空，未写入')
            return False

        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
        else:
            ws = self.wb.create_sheet(sheet_name)
        ws.freeze_panes = 'A2'
        if ws.max_row != 1:
            data = data[1:]
        for row in data:
            ws.append(row)
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, len(data[0]) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        for cell in ws[1]:
            if cell.value == '是否作弊':
                column_letter = cell.column_letter
                column_cells = ws[column_letter]
                for cell in column_cells[1:]:
                    if cell.value == '是':
                        cell.fill = styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        return True

    def save(self) -> bool:
        '''保存表格数据到本地'''     
        try:
            self.wb.active = 0
            self.wb.save(self.file_path)
            self.wb.close()
            return True
        except PermissionError as e:
            logging.error(f'文件保存失败!请勿在打开表格时操作：{e}')
        return False
